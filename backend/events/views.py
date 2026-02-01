from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.views import APIView
from rest_framework.permissions import IsAuthenticated, AllowAny
from django.shortcuts import get_object_or_404
from django.http import HttpResponse
from django.db.models import Q
from django.utils import timezone
from .models import Event, Incident, Investigation, ChartDefinition
from .serializers import EventSerializer, IncidentSerializer, InvestigationSerializer, ChartDefinitionSerializer
from core.models import AuditLog
from .services import promote_event_to_incident
from .sync import OnSpringSyncManager
from .permissions import HasAPIKey
import json
import logging
import io
from openpyxl import Workbook
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter

logger = logging.getLogger(__name__)

class OnSpringWebhookView(APIView):
    """
    Receives updates from OnSpring.
    """
    permission_classes = [HasAPIKey]

    def post(self, request):
        data = request.data
        os_id = data.get('recordId')

        if os_id:
            OnSpringSyncManager.handle_incoming_update(os_id, data)
            return Response({"status": "received"}, status=status.HTTP_200_OK)
        return Response({"error": "missing recordId"}, status=status.HTTP_400_BAD_REQUEST)

class GlobalSearchView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        query = request.query_params.get('q', '')
        if not query:
            return Response({"results": []})

        # Search Events
        events = Event.objects.filter(
            Q(title__icontains=query) |
            Q(id__icontains=query) |
            Q(source__icontains=query)
        )[:5]

        # Search Incidents
        incidents = Incident.objects.filter(
            Q(title__icontains=query) |
            Q(id__icontains=query)
        )[:5]

        # Search Investigations
        investigations = Investigation.objects.filter(
            Q(event__title__icontains=query) |
            Q(id__icontains=query)
        )[:5]

        results = []
        for e in events:
            results.append({"type": "Event", "id": e.id, "title": e.title, "link": "/events"})

        for i in incidents:
            results.append({"type": "Incident", "id": i.id, "title": i.title, "link": "/incidents"})

        for inv in investigations:
            results.append({"type": "Investigation", "id": inv.id, "title": inv.event.title, "link": f"/investigations/{inv.id}"})

        return Response({"results": results})

class EventViewSet(viewsets.ModelViewSet):
    queryset = Event.objects.all().order_by('-created_at')
    serializer_class = EventSerializer
    permission_classes = [IsAuthenticated]

    @action(detail=True, methods=['post'])
    def promote(self, request, pk=None):
        event = self.get_object()
        incident = promote_event_to_incident(event, request.user)
        return Response(IncidentSerializer(incident).data, status=status.HTTP_201_CREATED)

    @action(detail=False, methods=['post'])
    def bulk_status_update(self, request):
        ids = request.data.get('ids', [])
        new_status = request.data.get('status')
        reason = request.data.get('reason', '')

        if not ids or not new_status:
            return Response({"error": "Missing ids or status"}, status=400)

        events = Event.objects.filter(id__in=ids)
        updated_count = 0

        for event in events:
            if event.status == new_status:
                continue

            old_status = event.status
            event.status = new_status
            event.status_change_reason = reason
            event.save()

            # Audit Log
            AuditLog.objects.create(
                event=event,
                user=request.user,
                action=f"STATUS_CHANGE: {old_status} -> {new_status}",
                details={"reason": reason}
            )

            # Syslog
            log_entry = {
                "event": "status_change",
                "event_id": event.id,
                "old_status": old_status,
                "new_status": new_status,
                "user": request.user.username,
                "reason": reason
            }
            logger.info(json.dumps(log_entry))
            print(f"SYSLOG: {json.dumps(log_entry)}")

            # Investigation Creation Logic (if TP)
            if new_status == Event.Status.TRUE_POSITIVE:
                 inv, created = Investigation.objects.get_or_create(event=event)
                 if created:
                     inv.indicators.set(event.indicators.all())

            updated_count += 1

        return Response({"updated": updated_count})

class InvestigationViewSet(viewsets.ModelViewSet):
    queryset = Investigation.objects.all().order_by('-created_at')
    serializer_class = InvestigationSerializer
    permission_classes = [IsAuthenticated]

    @action(detail=True, methods=['post'])
    def promote(self, request, pk=None):
        investigation = self.get_object()
        event = investigation.event
        incident = promote_event_to_incident(event, request.user)
        return Response(IncidentSerializer(incident).data, status=status.HTTP_201_CREATED)

    @action(detail=True, methods=['post'])
    def add_timeline_event(self, request, pk=None):
        investigation = self.get_object()
        entry = request.data.get('entry')
        if entry:
             investigation.timeline.append({
                 "timestamp": str(timezone.now()), # Need timezone import
                 "entry": entry,
                 "user": request.user.username
             })
             investigation.save()
        return Response(investigation.timeline)

    @action(detail=True, methods=['post'])
    def exclude_indicator(self, request, pk=None):
        investigation = self.get_object()
        indicator_id = request.data.get('indicator_id')
        if indicator_id:
            investigation.indicators.remove(indicator_id)
        return Response({"status": "removed"})

    @action(detail=True, methods=['post'])
    def add_tag(self, request, pk=None):
        investigation = self.get_object()
        tag = request.data.get('tag')
        if tag and tag not in investigation.tags:
            investigation.tags.append(tag)
            investigation.save()
        return Response(investigation.tags)

    @action(detail=True, methods=['post'])
    def remove_tag(self, request, pk=None):
        investigation = self.get_object()
        tag = request.data.get('tag')
        if tag and tag in investigation.tags:
            investigation.tags.remove(tag)
            investigation.save()
        return Response(investigation.tags)

class ChartDefinitionViewSet(viewsets.ModelViewSet):
    serializer_class = ChartDefinitionSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        return ChartDefinition.objects.filter(user=self.request.user) | ChartDefinition.objects.filter(is_global=True)

    def perform_create(self, serializer):
        serializer.save(user=self.request.user)

class ReportingViewSet(viewsets.ViewSet):
    permission_classes = [IsAuthenticated]

    def _get_report_data(self, config):
        model_name = config.get('model', 'Event')
        group_by = config.get('group_by', 'source')

        if model_name == 'Event':
            qs = Event.objects.all()
        elif model_name == 'Incident':
            qs = Incident.objects.all()
        else:
            return []

        filters = config.get('filters', {})
        if filters:
             clean_filters = {k: v for k, v in filters.items() if v is not None}
             if clean_filters:
                 qs = qs.filter(**clean_filters)

        from django.db.models import Count
        data = qs.values(group_by).annotate(count=Count('id')).order_by('-count')

        return [
            {"name": str(item[group_by] or "Unknown"), "value": item['count']}
            for item in data
        ]

    @action(detail=False, methods=['post'])
    def generate(self, request):
        config = request.data.get('query_config', {})
        data = self._get_report_data(config)
        return Response(data)

    @action(detail=False, methods=['post'])
    def export_excel(self, request):
        data_payload = request.data.get('data', {})
        charts = data_payload.get('charts', [])

        wb = Workbook()
        if wb.sheetnames:
            wb.remove(wb.active)

        if not charts:
             ws = wb.create_sheet("Report")
             ws.append(["No charts selected"])

        for chart in charts:
             title = chart.get('title', 'Chart')
             safe_title = "".join(c for c in title if c.isalnum() or c in (' ', '_', '-'))[:30]
             ws = wb.create_sheet(title=safe_title)

             config = chart.get('query_config', {})
             report_data = self._get_report_data(config)

             ws.append(["Group", "Count"])
             for row in report_data:
                 ws.append([row['name'], row['value']])

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        response = HttpResponse(output.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="report.xlsx"'
        return response

    @action(detail=False, methods=['post'])
    def export_pdf(self, request):
        data_payload = request.data.get('data', {})
        charts = data_payload.get('charts', [])

        buffer = io.BytesIO()
        p = canvas.Canvas(buffer, pagesize=letter)
        p.drawString(100, 750, "Security Report")
        p.drawString(100, 730, f"Generated by: {request.user.username}")

        y = 700

        for chart in charts:
             if y < 100:
                 p.showPage()
                 y = 750

             title = chart.get('title', 'Chart')
             p.setFont("Helvetica-Bold", 12)
             p.drawString(100, y, title)
             y -= 20
             p.setFont("Helvetica", 10)

             config = chart.get('query_config', {})
             report_data = self._get_report_data(config)

             for row in report_data:
                 if y < 50:
                     p.showPage()
                     y = 750
                 p.drawString(120, y, f"{row['name']}: {row['value']}")
                 y -= 15

             y -= 20

        p.showPage()
        p.save()

        buffer.seek(0)
        response = HttpResponse(buffer.read(), content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="report.pdf"'
        return response

from integrations.ops.clients import WebExClient

class IncidentViewSet(viewsets.ModelViewSet):
    queryset = Incident.objects.all().order_by('-created_at')
    serializer_class = IncidentSerializer
    permission_classes = [IsAuthenticated]

    @action(detail=True, methods=['post'])
    def create_war_room(self, request, pk=None):
        incident = self.get_object()
        client = WebExClient(mock=True)
        room_url = client.create_room(f"War Room: {incident.title}", ["analyst@example.com"])
        return Response({"room_url": room_url}, status=status.HTTP_201_CREATED)

class WebhookIngestView(APIView):
    """
    Public endpoint for ingesting alerts from SIEMs.
    """
    permission_classes = [HasAPIKey]

    def post(self, request, source=None):
        data = request.data

        # Simple normalization logic
        title = data.get('title', 'Untitled Alert')
        description = data.get('description', json.dumps(data))

        # Detect source if not provided in URL
        if not source:
            source = Event.Source.OTHER
            if 'crowdstrike' in json.dumps(data).lower():
                source = Event.Source.CROWDSTRIKE

        event = Event.objects.create(
            source=source.upper(),
            title=title,
            description=description,
            raw_data=data,
            status=Event.Status.NEW
        )

        return Response(EventSerializer(event).data, status=status.HTTP_201_CREATED)
